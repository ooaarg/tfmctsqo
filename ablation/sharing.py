"""Convert runs/<id>/ ↔ multi-sheet xlsx workbooks."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
import lib
from sweep import luby_cap

BASELINE_CONFIGS = {"dp", "geqo"}
WINNER_TOL_COST = 0.01
WINNER_TOL_TIME = 0.05

PER_QUERY_METRIC_GROUPS = (("COST", "oracle_cost"), ("RUNTIME", "runtime_ms"), ("E2E", "e2e_ms"))

_AGG_SHORT = {"best": "best", "average": "avg"}


def _resolve_gucs(config_kind: str, row_gucs: dict, cfg_gucs: dict) -> dict:
    out: dict = dict(cfg_gucs.get("shared", {}) or {})
    out.update(cfg_gucs.get(config_kind, {}) or {})
    if isinstance(row_gucs, dict):
        flat = {k: v for k, v in row_gucs.items() if not isinstance(v, dict)}
        nested = row_gucs.get(config_kind) or {}
        out.update(flat)
        out.update(nested)

    if config_kind == "single" and "start_budget" not in out:
        luby_b = dict(cfg_gucs.get("shared", {}) or {})
        luby_b.update(cfg_gucs.get("luby", {}) or {})
        if isinstance(row_gucs, dict):
            luby_b.update({k: v for k, v in row_gucs.items() if not isinstance(v, dict)})
            luby_b.update(row_gucs.get("luby") or {})
        if "start_budget" in luby_b and "phases" in luby_b:
            out["start_budget"] = luby_cap(int(luby_b["start_budget"]), int(luby_b["phases"]))
    return out


def cell_to_mode_name(
    config: str, reward: str, agg: str, resolved_gucs: dict, *, suffix_with_reward_agg: bool
) -> str:
    if config in BASELINE_CONFIGS:
        return config

    if config == "luby":
        sb = resolved_gucs.get("start_budget", "?")
        base = f"mcts_budget_{sb}"
    elif config == "single":
        sb = resolved_gucs.get("start_budget", "?")
        d = resolved_gucs.get("depth", "?")
        base = f"fat_{sb}_d{d}"
    else:
        base = config

    if suffix_with_reward_agg:
        rw = lib._REWARD_SHORT.get(reward, str(reward)[:6])
        ag = _AGG_SHORT.get(agg, str(agg)[:4])
        return f"{base}__{rw}_{ag}"
    return base


_MODE_LUBY_RE = re.compile(r"^mcts_budget_(\d+)(?:_(plus_tail))?(?:__([^_]+)_([^_]+))?$")
_MODE_SINGLE_RE = re.compile(r"^fat_(\d+)_d(?:epth)?(\d+)(?:__([^_]+)_([^_]+))?$")


def mode_to_cell(mode: str) -> dict:
    base = {"reward": "neg_log", "agg": "best", "combo_id": mode}
    if mode in BASELINE_CONFIGS:
        return {**base, "config": mode, "reward": "n/a", "agg": "n/a", "combo_id": "", "gucs": {}}

    m = _MODE_LUBY_RE.match(mode)
    if m:
        sb = int(m.group(1))
        gucs: dict = {"start_budget": sb}
        if m.group(2) == "plus_tail":
            gucs["tail_after_phase3"] = "on"
        return {**base, "config": "luby", "gucs": gucs}

    m = _MODE_SINGLE_RE.match(mode)
    if m:
        sb, d = int(m.group(1)), int(m.group(2))
        return {**base, "config": "single", "gucs": {"start_budget": sb, "depth": d, "phases": 1}}

    return {**base, "config": "single", "gucs": {}}


def _winners_for_row(values_by_mode: dict[str, float], *, tol: float) -> list[str]:
    pairs = [(m, v) for m, v in values_by_mode.items() if v is not None and not pd.isna(v)]
    if not pairs:
        return []
    best = min(v for _, v in pairs)
    if best == 0:
        return [m for m, v in pairs if v == 0]
    return [m for m, v in pairs if v <= best * (1.0 + tol)]


def _safe_med(s: pd.Series) -> float | None:
    s = s.dropna()
    return None if s.empty else float(s.median())


def _safe_min(s: pd.Series) -> float | None:
    s = s.dropna()
    return None if s.empty else float(s.min())


def _safe_mean(s: pd.Series) -> float | None:
    s = s.dropna()
    return None if s.empty else float(s.mean())


def _enrich_dataframe(df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    df = df.copy()
    if "timed_out" not in df.columns:
        df["timed_out"] = False
    df["timed_out"] = df["timed_out"].fillna(False).astype(bool)
    if "combo_id" not in df.columns:
        df["combo_id"] = ""
    df["combo_id"] = df["combo_id"].fillna("").astype(str)
    if "gucs" not in df.columns:
        df["gucs"] = [{} for _ in range(len(df))]
    df["gucs"] = df["gucs"].apply(lambda x: x if isinstance(x, dict) else {})

    suffix = len(set(df["reward"]) - {"n/a"}) > 1 or len(set(df["agg"]) - {"n/a"}) > 1

    cfg_gucs = cfg.get("gucs", {}) or {}
    resolved = [_resolve_gucs(r["config"], r["gucs"], cfg_gucs) for _, r in df.iterrows()]
    df["resolved_gucs"] = resolved
    df["mode"] = [
        cell_to_mode_name(r["config"], r["reward"], r["agg"], rg, suffix_with_reward_agg=suffix)
        for (_, r), rg in zip(df.iterrows(), resolved, strict=False)
    ]

    df["runtime_ms"] = df.get("exec_time_ms")
    df["e2e_ms"] = df.get("plan_time_ms", pd.Series([0.0] * len(df))).fillna(0.0) + df.get(
        "exec_time_ms", pd.Series([0.0] * len(df))
    ).fillna(0.0)
    df["planning_overhead_ms"] = df.get("plan_time_ms")
    if "mcts_best_cost" not in df.columns:
        df["mcts_best_cost"] = pd.NA
    if "plan_cost" not in df.columns:
        df["plan_cost"] = pd.NA
    df["oracle_cost"] = df["mcts_best_cost"].fillna(df["plan_cost"])
    df["final_plan_cost"] = df["plan_cost"]
    return df


def _build_protocol_rows(
    run_dir: Path, cfg: dict, df: pd.DataFrame, modes: list[str], mode_to_gucs: dict[str, dict]
) -> list[tuple]:
    rows: list[tuple] = [("section", "field", "value")]
    created_at = ""
    if not df.empty and "ts" in df.columns:
        try:
            created_at = str(df["ts"].iloc[0])
        except Exception:
            created_at = ""
    rows += [
        ("Run", "run_id", run_dir.name),
        ("Run", "created_at", created_at),
        ("Run", "type", "exported from ablation run"),
        ("Runtime", "database", cfg.get("db", "")),
        ("Runtime", "pgbin", cfg.get("pgbin", "")),
        ("Runtime", "pgdata", cfg.get("pgdata", "")),
        ("Dataset", "query_count", len(cfg.get("queries", []) or [])),
        ("Dataset", "queries", ",".join(cfg.get("queries", []) or [])),
        ("Dataset", "modes", ",".join(modes)),
        ("Dataset", "repeat_count", str(cfg.get("n_seeds", ""))),
        ("Measurement", "primary_metric", "runtime_ms"),
        (
            "Measurement",
            "secondary_metrics",
            "e2e_ms, planning_overhead_ms, oracle_cost, final_plan_cost",
        ),
        ("Measurement", "cache_policy", "cold_drop_caches"),
        ("Measurement", "statement_timeout", _read_statement_timeout(run_dir)),
        ("MCTS", "reward_modes", ",".join(cfg.get("reward_modes", []) or [])),
        ("MCTS", "uct_aggregations", ",".join(cfg.get("uct_aggregations", []) or [])),
        ("MCTS", "seed_policy", f"seeds 1..{cfg.get('n_seeds', '?')}"),
    ]
    for mode in modes:
        gucs = mode_to_gucs.get(mode, {})
        if not gucs:
            continue
        bundle = "; ".join(f"{k}={v}" for k, v in sorted(gucs.items()))
        rows.append(("MCTS Budgets", mode, bundle))
    return rows


def _read_statement_timeout(run_dir: Path) -> str:
    for cand in ("luby.sql", "single.sql", "dp.sql", "geqo.sql"):
        p = run_dir / cand
        if not p.exists():
            continue
        m = re.search(r"statement_timeout\s*=\s*'?([^';\n]+)'?", p.read_text())
        if m:
            return m.group(1).strip()
    return ""


def _build_runtime_summary(df: pd.DataFrame, modes: list[str]) -> list[list]:
    header = [
        "query",
        "alias_count",
        "mode",
        "ok_runs",
        "status",
        "oracle_cost",
        "final_plan_cost",
        "runtime_ms",
        "e2e_ms",
        "planning_overhead_ms",
    ]
    out: list[list] = [header]
    queries = sorted(df["query"].unique()) if not df.empty else []
    for q in queries:
        sub_q = df[df["query"] == q]
        try:
            alias = int(sub_q["rels"].dropna().iloc[0])
        except (IndexError, ValueError, KeyError):
            alias = None
        for mode in modes:
            sub = sub_q[sub_q["mode"] == mode]
            if sub.empty:
                continue
            n = len(sub)
            n_to = int(sub["timed_out"].sum())
            ok = n - n_to
            status = "ok" if n_to == 0 else ("partial" if ok > 0 else "failed")
            row = [
                q,
                alias,
                mode,
                ok,
                status,
                _safe_min(sub["oracle_cost"]),
                _safe_min(sub["final_plan_cost"]),
                _safe_med(sub["runtime_ms"]),
                _safe_med(sub["e2e_ms"]),
                _safe_med(sub["planning_overhead_ms"]),
            ]
            out.append(row)
    return out


def _per_query_metric_value(sub: pd.DataFrame, metric: str) -> float | None:
    if metric == "oracle_cost":
        return _safe_min(sub[metric])
    return _safe_med(sub[metric])


def _build_per_query(
    df: pd.DataFrame, modes: list[str], *, normalize_to: str | None = None
) -> list[list]:
    grid: dict[tuple[str, str], dict[str, float | None]] = {}
    for (q, mode), sub in df.groupby(["query", "mode"]):
        grid[(q, mode)] = {
            metric: _per_query_metric_value(sub, metric) for _, metric in PER_QUERY_METRIC_GROUPS
        }

    top: list = ["QUERY INFO", None, None, "WINNERS", None, None]
    info_header: list = ["suite", "query", "numrels", "COST", "RUNTIME", "E2E"]
    for group_name, _ in PER_QUERY_METRIC_GROUPS:
        title = group_name if normalize_to is None else f"{group_name} / {normalize_to}"
        top.append(title)
        top.extend([None] * (len(modes) - 1))
        info_header.extend(modes)
    out: list[list] = [top, info_header]

    queries = sorted(df["query"].unique()) if not df.empty else []
    for q in queries:
        sub_q = df[df["query"] == q]
        try:
            suite = str(sub_q["source"].dropna().iloc[0])
        except (IndexError, KeyError):
            suite = ""
        try:
            numrels = int(sub_q["rels"].dropna().iloc[0])
        except (IndexError, ValueError, KeyError):
            numrels = None

        # Winners per metric (against raw, un-normalized values).
        winners: dict[str, list[str]] = {}
        for _group, metric in PER_QUERY_METRIC_GROUPS:
            tol = WINNER_TOL_COST if metric == "oracle_cost" else WINNER_TOL_TIME
            vals = {mode: grid.get((q, mode), {}).get(metric) for mode in modes}
            winners[metric] = _winners_for_row(vals, tol=tol)

        row: list = [
            suite,
            q,
            numrels,
            ";".join(winners["oracle_cost"]),
            ";".join(winners["runtime_ms"]),
            ";".join(winners["e2e_ms"]),
        ]

        denom: dict[str, float | None] = {}
        if normalize_to is not None:
            denom = {
                metric: grid.get((q, normalize_to), {}).get(metric)
                for _, metric in PER_QUERY_METRIC_GROUPS
            }

        for _, metric in PER_QUERY_METRIC_GROUPS:
            for mode in modes:
                v = grid.get((q, mode), {}).get(metric)
                if normalize_to is not None and v is not None and denom.get(metric):
                    v = v / denom[metric]
                row.append(v)
        out.append(row)
    return out


def _build_aggregate(per_query_rows: list[list], df: pd.DataFrame, modes: list[str]) -> list[list]:
    header = [
        "algorithm",
        "completed_count",
        "failed_count",
        "skipped_count",
        "runtime_winner_count",
        "e2e_winner_count",
        "cost_winner_count",
        "oracle_pick_count",
        "avg_runtime_ms",
        "median_runtime_ms",
        "avg_e2e_ms",
        "median_e2e_ms",
        "avg_planning_overhead_ms",
        "median_planning_overhead_ms",
        "best_query",
        "worst_query",
        "notes",
    ]
    out: list[list] = [header]

    # per_query_rows: [top_header, info_header, *data_rows]; cols 3,4,5 hold
    # semicolon-joined cost/runtime/e2e winners.
    win_cost: dict[str, int] = defaultdict(int)
    win_run: dict[str, int] = defaultdict(int)
    win_e2e: dict[str, int] = defaultdict(int)
    oracle_pick: dict[str, int] = defaultdict(int)
    for r in per_query_rows[2:]:
        for m in (r[3] or "").split(";"):
            if m:
                win_cost[m] += 1
                oracle_pick[m] += 1
        for m in (r[4] or "").split(";"):
            if m:
                win_run[m] += 1
        for m in (r[5] or "").split(";"):
            if m:
                win_e2e[m] += 1

    note = "Generated by ablation/sharing.py export_workbook."
    grand_total_rows = 0
    for mode in modes:
        sub = df[df["mode"] == mode]
        if sub.empty:
            out.append(
                [mode, 0, 0, 0, 0, 0, 0, 0, None, None, None, None, None, None, "", "", note]
            )
            continue
        completed = int((~sub["timed_out"]).sum())
        failed = int(sub["timed_out"].sum())
        grand_total_rows += len(sub)
        per_q = sub.groupby("query")["runtime_ms"].median().dropna()
        best_q = per_q.idxmin() if not per_q.empty else ""
        worst_q = per_q.idxmax() if not per_q.empty else ""
        out.append(
            [
                mode,
                completed,
                failed,
                0,
                win_run.get(mode, 0),
                win_e2e.get(mode, 0),
                win_cost.get(mode, 0),
                oracle_pick.get(mode, 0),
                _safe_mean(sub["runtime_ms"]),
                _safe_med(sub["runtime_ms"]),
                _safe_mean(sub["e2e_ms"]),
                _safe_med(sub["e2e_ms"]),
                _safe_mean(sub["planning_overhead_ms"]),
                _safe_med(sub["planning_overhead_ms"]),
                best_q,
                worst_q,
                note,
            ]
        )

    n_queries = len(set(df["query"])) if not df.empty else 0
    out.append(
        [
            "TOTAL",
            int((~df["timed_out"]).sum()) if not df.empty else 0,
            int(df["timed_out"].sum()) if not df.empty else 0,
            0,
            sum(win_run.values()),
            sum(win_e2e.values()),
            sum(win_cost.values()),
            sum(oracle_pick.values()),
            None,
            None,
            None,
            None,
            None,
            None,
            "",
            "",
            f"{n_queries} queries; {grand_total_rows} measured rows.",
        ]
    )
    return out


def export_workbook(run_dir: Path, out_path: Path) -> Path:
    import openpyxl

    run_dir = Path(run_dir)
    cfg_path = run_dir / "config.json"
    cfg = json.loads(cfg_path.read_text()) if cfg_path.exists() else {}

    df_raw = lib.load_jsonl(run_dir / "results.jsonl")
    df = _enrich_dataframe(df_raw, cfg) if not df_raw.empty else df_raw

    modes_seen = list(dict.fromkeys(df["mode"].tolist())) if not df.empty else []
    base_modes = [m for m in ("dp", "geqo") if m in modes_seen]
    mcts_modes = sorted(m for m in modes_seen if m not in BASELINE_CONFIGS)
    modes = base_modes + mcts_modes

    mode_to_gucs: dict[str, dict] = {}
    if not df.empty:
        for mode in modes:
            sub = df[df["mode"] == mode]
            if sub.empty:
                continue
            mode_to_gucs[mode] = sub["resolved_gucs"].iloc[0]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("protocol")
    for row in _build_protocol_rows(run_dir, cfg, df, modes, mode_to_gucs):
        ws.append(list(row))

    ws = wb.create_sheet("runtime_summary")
    for row in _build_runtime_summary(df, modes):
        ws.append(row)

    pq_rows = _build_per_query(df, modes)
    ws = wb.create_sheet("per_query")
    for row in pq_rows:
        ws.append(row)

    baseline = next((m for m in ("geqo", "dp") if m in modes), None)
    if baseline is not None:
        pqr_rows = _build_per_query(df, modes, normalize_to=baseline)
        ws = wb.create_sheet("per_query_relative")
        for row in pqr_rows:
            ws.append(row)

    ws = wb.create_sheet("aggregate")
    for row in _build_aggregate(pq_rows, df, modes):
        ws.append(row)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXPDATA_DIR = REPO_ROOT / "expdata"

# Top-level files always worth committing.  `pid` is excluded (process artifact).
_ALWAYS_TOP_FILES = ("status", "README.md", "config.json", "results.jsonl", "errors.log")


def export_expdata(
    run_dir: Path,
    dest_root: Path | None = None,
    *,
    slim: bool = False,
    overwrite: bool = False,
) -> dict:
    """Copy `run_dir` into `<dest_root>/<run_name>/` for git committing.

    slim=True drops the per-(cell, query) `.plan` directories (~95% of the size)
    and keeps only results.jsonl + metadata, which is all `app.py` needs to
    re-render aggregations.

    Returns: {"path": Path, "n_files": int, "size_bytes": int, "slim": bool, "skipped_plans": int}.
    """
    run_dir = Path(run_dir).resolve()
    if not run_dir.is_dir():
        raise FileNotFoundError(f"run_dir does not exist: {run_dir}")
    dest_root = (Path(dest_root) if dest_root is not None else DEFAULT_EXPDATA_DIR).resolve()
    dest = (dest_root / run_dir.name).resolve()
    # Guard against a caller passing a weird dest_root that resolves outside itself.
    if not dest.is_relative_to(dest_root):
        raise ValueError(f"dest {dest} is not under dest_root {dest_root}")
    if dest.exists():
        if not overwrite:
            raise FileExistsError(
                f"{dest} already exists — pass overwrite=True to replace it."
            )
        shutil.rmtree(dest)
    dest.mkdir(parents=True, exist_ok=True)

    n_files = 0
    size_bytes = 0
    skipped_plans = 0

    def _copy_file(src: Path, dst: Path) -> None:
        nonlocal n_files, size_bytes
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dst)
        n_files += 1
        size_bytes += dst.stat().st_size

    for name in _ALWAYS_TOP_FILES:
        src = run_dir / name
        if src.is_file():
            _copy_file(src, dest / name)
    for sql in run_dir.glob("*.sql"):
        _copy_file(sql, dest / sql.name)

    for child in run_dir.iterdir():
        if not child.is_dir():
            continue
        # sweep.py only writes <cell>/<query>/seed-N.plan beneath run_dir.
        plans = list(child.glob("*/*.plan"))
        if slim:
            skipped_plans += len(plans)
            continue
        for f in plans:
            rel = f.relative_to(run_dir)
            _copy_file(f, dest / rel)

    return {
        "path": dest,
        "n_files": n_files,
        "size_bytes": size_bytes,
        "slim": slim,
        "skipped_plans": skipped_plans,
    }


def _row_value(idx: dict, row: list, name: str):
    i = idx.get(name)
    if i is None or i >= len(row):
        return None
    return row[i]


def _read_sheet(ws) -> list[list]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def import_workbook(xlsx_path: Path, out_dir: Path) -> Path:
    import openpyxl

    xlsx_path = Path(xlsx_path)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheets = set(wb.sheetnames)

    proto_kv: dict[tuple[str, str], str] = {}
    mode_bundles: dict[str, dict] = {}
    if "protocol" in sheets:
        rows = _read_sheet(wb["protocol"])
        for r in rows[1:]:
            if len(r) < 3:
                continue
            section, field, value = r[0], r[1], r[2]
            if section is None or field is None:
                continue
            proto_kv[(str(section), str(field))] = "" if value is None else str(value)
            if section == "MCTS Budgets" and value:
                # Parse "start_budget=80; phases=5; depth=2; ..." into dict.
                bundle: dict = {}
                for pair in str(value).split(";"):
                    pair = pair.strip()
                    if "=" not in pair:
                        continue
                    k, _, v = pair.partition("=")
                    k, v = k.strip(), v.strip()
                    try:
                        bundle[k] = int(v)
                    except ValueError:
                        try:
                            bundle[k] = float(v)
                        except ValueError:
                            bundle[k] = v
                mode_bundles[str(field)] = bundle

    queries_csv = proto_kv.get(("Dataset", "queries"), "")
    queries = [q.strip() for q in queries_csv.split(",") if q.strip()]
    modes_csv = proto_kv.get(("Dataset", "modes"), "")
    modes = [m.strip() for m in modes_csv.split(",") if m.strip()]
    try:
        n_seeds = int(proto_kv.get(("Dataset", "repeat_count"), "1"))
    except ValueError:
        n_seeds = 1

    # ----- runtime_summary → results.jsonl -----
    if "runtime_summary" not in sheets:
        raise ValueError(f"{xlsx_path}: missing 'runtime_summary' sheet")
    rs_rows = _read_sheet(wb["runtime_summary"])
    header = rs_rows[0]
    idx = {name: i for i, name in enumerate(header)}

    out_jsonl = out_dir / "results.jsonl"
    n_rows_written = 0
    with out_jsonl.open("w") as f:
        for r in rs_rows[1:]:
            mode = _row_value(idx, r, "mode")
            if not mode:
                continue
            cell = mode_to_cell(str(mode))
            for k, v in mode_bundles.get(mode, {}).items():
                if k in (
                    "reward_mode",
                    "restart_policy",
                    "depth_policy",
                    "tail_after_phase3",
                    "luby_coefficients",
                    "phase_budgets",
                    "total_nominal_iterations",
                ):
                    continue
                cell["gucs"].setdefault(k, v)
            row = {
                "ts": dt.datetime.now().isoformat(timespec="milliseconds") + "Z",
                "query": _row_value(idx, r, "query"),
                "source": "",
                "config": cell["config"],
                "reward": cell["reward"],
                "agg": cell["agg"],
                "combo_id": cell["combo_id"],
                "gucs": cell["gucs"],
                "seed": 1,
                "plan_cost": _row_value(idx, r, "final_plan_cost"),
                "mcts_best_cost": _row_value(idx, r, "oracle_cost"),
                "plan_time_ms": _row_value(idx, r, "planning_overhead_ms"),
                "exec_time_ms": _row_value(idx, r, "runtime_ms"),
                "rels": _row_value(idx, r, "alias_count"),
                "n_seeds_collapsed": _row_value(idx, r, "ok_runs"),
                "imported_status": _row_value(idx, r, "status"),
            }
            row = {k: v for k, v in row.items() if v is not None}
            f.write(json.dumps(row, default=str) + "\n")
            n_rows_written += 1

    cfg = {
        "name": out_dir.name,
        "description": f"Imported from {xlsx_path.name}.",
        "queries": queries,
        "reward_modes": [
            m.strip()
            for m in proto_kv.get(("MCTS", "reward_modes"), "neg_log").split(",")
            if m.strip()
        ]
        or ["neg_log"],
        "uct_aggregations": [
            m.strip()
            for m in proto_kv.get(("MCTS", "uct_aggregations"), "best").split(",")
            if m.strip()
        ]
        or ["best"],
        "configs": _infer_configs_from_modes(modes),
        "n_seeds": n_seeds,
        "n_combos": max(
            1, len({mode_to_cell(m)["combo_id"] for m in modes if m not in BASELINE_CONFIGS})
        ),
        "grid_varying": [],
        "gucs": {"shared": {}, "luby": {}, "single": {}},
        "sources": [],
        "db": proto_kv.get(("Runtime", "database"), ""),
        "pgbin": proto_kv.get(("Runtime", "pgbin"), ""),
        "pgdata": proto_kv.get(("Runtime", "pgdata"), ""),
        "imported_from": str(xlsx_path),
    }
    (out_dir / "config.json").write_text(json.dumps(cfg, indent=2, default=str))

    (out_dir / "README.md").write_text(
        f"# {out_dir.name}\n\n"
        f"Imported from `{xlsx_path}` via `ablation/sharing.py import`.\n\n"
        f"- queries: {len(queries)}\n"
        f"- modes: {len(modes)}\n"
        f"- rows written: {n_rows_written}\n"
        f"\n*Plan files are not available for imported runs.*\n"
    )
    lib.write_status(out_dir, lib.STATUS_DONE)
    return out_dir


def _infer_configs_from_modes(modes: list[str]) -> list[str]:
    out = []
    for m in modes:
        if m in BASELINE_CONFIGS and m not in out:
            out.append(m)
        elif _MODE_LUBY_RE.match(m) and "luby" not in out:
            out.append("luby")
        elif _MODE_SINGLE_RE.match(m) and "single" not in out:
            out.append("single")
    return out


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    sub = p.add_subparsers(dest="cmd", required=True)

    pe = sub.add_parser("export", help="Convert runs/<id>/ → xlsx workbook.")
    pe.add_argument("run_dir", type=Path)
    pe.add_argument("-o", "--out", type=Path, default=None)

    pi = sub.add_parser("import", help="Convert xlsx workbook → runs/<id>/.")
    pi.add_argument("xlsx", type=Path)
    pi.add_argument("-o", "--out", type=Path, default=None)

    pd_ = sub.add_parser("expdata", help="Copy runs/<id>/ into git-shareable expdata/<id>/.")
    pd_.add_argument("run_dir", type=Path)
    pd_.add_argument("-o", "--out", type=Path, default=None, help="dest root (default: ./expdata)")
    pd_.add_argument("--slim", action="store_true", help="skip *.plan files")
    pd_.add_argument("--overwrite", action="store_true", help="replace existing destination")

    args = p.parse_args(argv)

    if args.cmd == "export":
        run_dir = args.run_dir
        out = args.out or run_dir.with_suffix(".xlsx")
        path = export_workbook(run_dir, out)
        print(f"Wrote {path}")
        return 0

    if args.cmd == "import":
        xlsx = args.xlsx
        # Strip parenthesised suffixes / spaces; build a clean run-dir name.
        stem = re.sub(r"\s+\(\d+\)", "", xlsx.stem).strip().replace(" ", "_")
        out = args.out or Path("runs") / f"imported-{stem}"
        path = import_workbook(xlsx, out)
        print(f"Wrote {path}")
        return 0

    if args.cmd == "expdata":
        info = export_expdata(args.run_dir, args.out, slim=args.slim, overwrite=args.overwrite)
        kb = info["size_bytes"] / 1024
        print(f"Wrote {info['path']} ({info['n_files']} files, {kb:.1f} KiB)")
        if info["slim"] and info["skipped_plans"]:
            print(f"  (slim: skipped {info['skipped_plans']} .plan files)")
        return 0

    return 2


if __name__ == "__main__":
    sys.exit(main())
